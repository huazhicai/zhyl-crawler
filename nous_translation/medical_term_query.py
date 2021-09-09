# -*- coding:utf-8 -*-
from openpyxl import load_workbook

ENGLISH_2_CHINESE = {}
MEDICAL_FILE = 'file/常用临床医学名词（2019年版）.xlsx'


def load_medical_term(filename):
    wb = load_workbook(filename)
    ws = wb['Sheet1']
    for row in ws.iter_rows(min_row=2, max_row=42473, min_col=3, max_col=4):
        if row[-1].value:
            en = row[-1].value.strip().lower()
            zh = row[0].value.strip()
            ENGLISH_2_CHINESE[en] = zh


def medical_query(text):
    if not ENGLISH_2_CHINESE:
        load_medical_term(MEDICAL_FILE)
    return ENGLISH_2_CHINESE.get(text.strip().lower())


if __name__ == '__main__':
    print(medical_query('AIDS retinopathy'))