# -*- coding:utf-8 -*-
from openpyxl import load_workbook
import re


"""
yyyyMMdd
yyyy-MM-dd
yyyy/MM/dd
yyyy.MM.dd
yyyy年MM月dd日
yyyy年
yyyy-dd-yyy-dd月
"""

_formats = {
    'y-m-d': '{d.year}-{d.month}-{d.day}',
    'y/m/d': '{d.year}/{d.month}/{d.day}',
    'nyr': '{d.year}年{d.month}月{d.day}日'
}


class Date(object):
    def __init__(self, year, month, day):
        self.year = year
        self.month = month
        self.day = day

    def __format__(self, code):
        if code == '':
            code = 'y-m-d'
        fmt = _formats[code]
        return fmt.format(d=self)


def iter_rows_excel(filename, sheet=None, min_row=1, max_row=100, min_col=1, max_col=5):
    wb = load_workbook(filename)
    sheet_title = sheet or wb.get_sheet_names()[0]
    ws = wb[sheet_title]
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        yield row


def write_rows_excel(filename, rows, sheet=None):
    wb = load_workbook(filename)
    sheet_title = sheet or wb.get_sheet_names()[0]
    try:
        ws = wb[sheet_title]
    except KeyError:
        wb.create_sheet(sheet)
        ws = wb[sheet_title]

    for row in rows:
        ws.append(row)
    wb.save(filename)


def re_date(text):
    year, month, day = '(\d{4})', '(\d{1,2})', '(\d{1,2})'
    result = re.search(f'{year}-{month}-{day}', text)
    if result:
        date = result.group()
        return text.replace(date, ''), date
    result = re.search(f'{year}年{month}月{day}日|{year}年{month}月|{year}年', text)
    if result:
        date = result.group()
        return text.replace(date, ''), date.replace('年', '-').replace('月', '-').replace('日', '-').strip('-')
    result = re.search(r'(\d{4})\.(\d{1,2})\.(\d{1,2})', text)
    if result:
        date = result.group()
        return text.replace(date, ''), date.replace('.', '-')
    result = re.search(r'\d{4}/\d{1,2}/\d{1,2}', text)
    if result:
        date = result.group()
        return text.replace(date, ''), date.replace('/', '-')
    result = re.search(r'\d{4}\.\d{1,2}-\d{4}\.\d{1,2}', text)
    if result:
        date = result.group()
        return text.replace(date, ''), date.replace('.', '-')


def main():
    container = []
    for row in iter_rows_excel(file, sheet='附表1-2018年1月至2021年9月序时账 (3)', min_row=1, max_row=4034, min_col=8, max_col=8):
        new_row = []
        value = row[0].value
        new_row.append(value)
        result = re.search(r'\d{4}', value)
        if result and re_date(value):
            new_row.extend(re_date(value))
        # print(new_row)
        container.append(new_row)
    print(len(container))
    write_rows_excel(file, container, sheet='new')


if __name__ == '__main__':
    file = 'sample.xlsx'
    main()
