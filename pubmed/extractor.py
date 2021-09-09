# -*- coding:utf-8 -*-
"""
1、遇到大坑多线程传入列表参数，结果大部分线程取的数据都一样
2、1万条数据一次写入的速度是，一条条写入的 33.78倍，所以大数据应该多条写入
3、 线程超过50后：internal buffer error : Memory allocation failed : growing buffer
because 每个页面1万个， 50个页面50万个，还有其他的很多
"""
from retrying import retry
import requests
import logging
import csv
import threading
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from lxml.etree import XML

from config import title_2_factor

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(threadName)s - %(levelname)s - %(message)s')


def getLogger(module_name):
    return logging.getLogger(module_name)


logger = getLogger(threading.current_thread().name)

DEFAULT_FACTOR = '0.000'
GROUP_COUNT = 10000
HAS_DONE = False
MAX_THREAD_NUM = 36  # 超过40 本机内存不够
FAILED_PMID_QUEUE = []

SAMPLE_FILE = 'samples.csv'

LOCK = threading.Lock()


def _request(url, method='GET', *args):
    pass


@retry(stop_max_attempt_number=5, wait_random_min=1000, wait_random_max=3000)
def get_request(url):
    # logger.info(f'get_request: {url}')
    response = requests.get(url, timeout=10)
    if response.status_code != 200:
        raise Exception('Unexpected status_code %s' % response.status_code)
    return response.text


@retry(stop_max_attempt_number=5, wait_random_min=1000, wait_random_max=3000)
def post_request(url, group_ids, retmax):
    data = {'maximum': retmax, 'retmax': retmax, 'id': group_ids,
            'db': 'pubmed', 'retmode': 'xml'}
    # logger.info('post start %s' % len(pmid_list))
    # logger.info(pmid_list)
    response = requests.post(url, data=data, timeout=20)  # get 最大 342
    if response.status_code != 200:
        raise Exception('unexpected status_code %s ' % response.status_code)
    # logger.info('post success')
    return response.text


def get_one_page(url, group_ids, retmax=10000):
    try:
        return post_request(url, group_ids, retmax)
    except Exception as e:
        logger.error(f'{url}\n{e}')
        return


def read_pmid(file_csv):
    saved_pmid = read_sample()
    pmid_list = []
    with open(file_csv, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if not saved_pmid.get(row[0]):
                pmid_list.append(int(row[0]))
                if len(pmid_list) == GROUP_COUNT:
                    yield pmid_list
                    pmid_list.clear()
        yield pmid_list
        if not pmid_list:
            global HAS_DONE
            HAS_DONE = True


def read_sample():
    saved_pmid = {}
    with open(SAMPLE_FILE, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            saved_pmid[row[0]] = int(row[0])
    return saved_pmid


def parser(xml_tree):
    doc = XML(xml_tree)
    logger.info('return quantity %s' % len(doc.xpath('/PubmedArticleSet/PubmedArticle')))
    parser_data = []
    for branch in doc.xpath('/PubmedArticleSet/PubmedArticle'):
        pmid = branch.xpath('./MedlineCitation/PMID/text()')
        title = branch.xpath('.//Journal/Title/text()')
        publication = branch.xpath('.//PublicationTypeList//text()')
        pmid.extend(title)
        pmid.append('\n'.join(publication).strip())
        pmid.append(get_factor(title[0].lower()))
        parser_data.append(pmid)
    return parser_data


def get_factor(title):
    if title_2_factor.get(title):
        return title_2_factor.get(title)
    for full_title in title_2_factor.keys():
        if title in full_title or full_title in title:
            return title_2_factor.get(full_title)
    return DEFAULT_FACTOR


def read_impact_factor(filename):
    wb = load_workbook(filename)
    ws = wb['DJR_583']
    global title_2_factor
    title_2_factor = {}
    for row in ws.iter_rows(min_row=4, max_row=13013, max_col=5, min_col=2):
        factor = row[-1].value.strip()
        if factor == 'Not Available':
            factor = DEFAULT_FACTOR
        title_2_factor[row[0].value.lower()] = factor
    print(title_2_factor)


def write_sample(rows):
    with LOCK:
        with open(SAMPLE_FILE, 'a', newline='') as f:
            writer = csv.writer(f, dialect='excel')
            writer.writerows(rows)


def save_post_fail(pmid):
    FAILED_PMID_QUEUE.append(pmid)
    # with LOCK:
    #     with open('failure.csv', 'a', newline='') as f:
    #         writer = csv.writer(f, dialect='excel')
    #         writer.writerow(pmid)


def extractor_group(group_ids):
    global group_num
    xml_tree = get_one_page(post_url, group_ids, retmax=GROUP_COUNT)
    if xml_tree:
        write_sample(parser(xml_tree))
        # for row in parser(xml_tree):
        #     write_sample(row)
        group_num += 1
        logger.info('save group %s' % group_num)
    else:
        save_post_fail(group_ids)


def main():
    global group_num
    group_num = 0
    cycle_num = 0
    while not HAS_DONE:
        with ThreadPoolExecutor(MAX_THREAD_NUM) as executor:
            for group_pmid in read_pmid(pmid_file):
                print(group_pmid)
                group_ids = str(group_pmid)[1:-1]
                executor.submit(extractor_group, group_ids)

        logger.info('finish pmid')
        if FAILED_PMID_QUEUE:
            with ThreadPoolExecutor(MAX_THREAD_NUM) as executor:
                while FAILED_PMID_QUEUE:
                     executor.submit(extractor_group, FAILED_PMID_QUEUE[0])
                     FAILED_PMID_QUEUE.pop(0)
        cycle_num += 1
        FAILED_PMID_QUEUE.clear()
        logger.info('finish cycle: %s' % cycle_num)
    logger.info('***************finished all!*****************')


if __name__ == '__main__':
    pmid_file = 'semmedVER43_2021_R_CITATIONS.23871.csv'
    factor_file = 'bak/Journal-Impact-factor_2021.xlsx'
    post_url = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi'
    # read_impact_factor(factor_file)
    main()
