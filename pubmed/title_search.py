# -*- coding:utf-8 -*-
from collections import Counter

from openpyxl import load_workbook
from retrying import retry
import requests
import logging
import threading
import json
from extractor import get_request, parser
from concurrent.futures import ThreadPoolExecutor
from remedy import write_rows_csv
from lxml import etree

logging.basicConfig(level=logging.INFO, format='%(asctime)s-%(threadName)s-%(levelname)s- %(message)s')
logger = logging.getLogger(threading.current_thread().name)
TITLE_2_FACTOR = {}
DATA_CONTAINER = []
COMPARE = []
URL = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?db=pubmed&id={}&retmode=xml'
_LOCK = threading.Lock()


@retry(stop_max_attempt_number=4, wait_random_min=1000, wait_random_max=3000)
def get_request(url):
    logger.info(url)
    response = requests.get(url, timeout=10)
    if response.status_code != 200:
        raise Exception('unexpected status_code %s ' % response.status_code)
    # logger.info('post success')
    return response.text


def get_one_page(url):
    try:
        return get_request(url)
    except Exception as e:
        logger.error(f'{url}\n{e}')
        return


def parser_pmid(html):
    doc_tree = etree.HTML(html)
    pmid = doc_tree.xpath('//article/div[2]/div[1]/div[1]/span[5]/span/text()')
    link = URL.format(','.join(pmid))
    response = requests.get(link)
    parser_data = parser(response.text)
    titles = [row[1] for row in parser_data]
    counter = Counter(titles)
    title_list = [title for count, title in sorted(zip(counter.values(), counter.keys()))]
    return title_list  # titles


def iter_term_title(filename):
    wb = load_workbook(filename)
    ws = wb['DJR_583']
    seen = set()
    for row in ws.iter_rows(min_row=4, max_row=13013, max_col=6, min_col=1):
        title = row[1].value.strip()
        if title not in seen:
            seen.add(title)
            yield row


def search_title(row):
    """
    :param row: tuple(rank, title, '', cites, factor, score) Cell(ele)
    :return:
    """
    new_row = [i.value for i in row if i.value]
    title = new_row[1].strip()
    term = '+'.join(title.split(' '))
    html = get_one_page(url.format(term))
    if html:
        pmid_list = parser_pmid(html)
        new_row.extend(pmid_list)
        for title in pmid_list:
            TITLE_2_FACTOR[title.lower()] = new_row[4]
            new_row.append(title)
        pass
    COMPARE.append(new_row)
    with _LOCK:
        DATA_CONTAINER.append(new_row)


def main():
    with ThreadPoolExecutor(thread_num) as executor:
        for row in iter_term_title(filename):
            # print(row)
            executor.submit(search_title, row)
    DATA_CONTAINER.sort(key=lambda x: x[0])
    write_rows_csv(save_as_journal, DATA_CONTAINER)
    print('\n' * 5)
    print(TITLE_2_FACTOR)
    with open('bak/title.json', 'w') as f:
        json.dump(TITLE_2_FACTOR, f, indent=4)


if __name__ == '__main__':
    url = 'https://pubmed.ncbi.nlm.nih.gov/?term={}%5Bjournal%5D'
    filename = 'bak/Journal-Impact-factor_2021.xlsx'
    save_as_journal = 'thread_impact.csv'
    thread_num = 33
    main()
