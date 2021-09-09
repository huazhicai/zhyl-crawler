from remedy import csv_reader, write_rows_csv
from openpyxl import load_workbook

CONTAINER = []


def read_impact_factor(filename):
    wb = load_workbook(filename)
    ws = wb['DJR_583']
    for row in ws.iter_rows(min_row=4, max_row=13013, max_col=6, min_col=1):
        yield [str(i.value).strip() for i in row if i.value]
        # yield row
        # factor = row[-1].value.strip()


def start():
    seen = set()
    for row in csv_reader('journal_impact.csv'):
        if row[1] not in seen:
            seen.add(row[1].strip())
            CONTAINER.append(row)

    for row in read_impact_factor('bak/Journal-Impact-factor_2021.xlsx'):
        if row[1] not in seen:
            seen.add(row[1])
            CONTAINER.append(row)
    # CONTAINER.sort(key=lambda x: str(x[0]))
    write_rows_csv('mark.csv', CONTAINER)


if __name__ == '__main__':
    impact_file = 'bak/Journal-Impact-factor_2021.xlsx'
    start()
