# -*- coding:utf-8 -*-
import csv
import json
from openpyxl import load_workbook


def csv_reader(file):
    try:
        with open(file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                yield row
    except UnicodeDecodeError:
        with open(file, 'r') as f:
            reader = csv.reader(f)
            for row in reader:
                yield row


def write_row_csv(filename, row, mode='a', newline=''):
    with open(filename, mode, newline=newline, encoding='utf-8') as f:
        writer = csv.writer(f, dialect='excel')
        writer.writerow(row)


def write_rows_csv(filename, rows, mode='a', newline=''):
    with open(filename, mode, newline=newline, encoding='utf-8') as f:
        writer = csv.writer(f, dialect='excel')
        writer.writerows(rows)


def iter_rows_excel(filename, sheet=None, min_row=1, max_row=100, min_col=1, max_col=5):
    wb = load_workbook(filename)
    sheet_title = sheet or wb.get_sheet_names()[0]
    ws = wb[sheet_title]
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        yield row


def write_rows_excel(filename, rows, sheet=None):
    wb = load_workbook(filename)
    sheet_title = sheet or wb.get_sheet_names()[0]
    ws = wb[sheet_title]
    ws.append(rows)
    wb.save(filename)


def load_json(file):
    try:
        with open(file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except UnicodeDecodeError:
        with open(file, encoding='gbk') as f:
            return json.load(f)


def dump_json(file, doc):
    with open(file, 'w', encoding='utf-8') as f:
        json.dump(f, doc, ensure_ascii=False)


def reader_file(filename):
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            return f.read()
    except UnicodeDecodeError:
        with open(filename, 'r', encoding='gbk') as f:
            return f.read()


def iter_line(filename):
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            line_list = f.readlines()
    except UnicodeDecodeError:
        with open(filename, 'r', encoding='gbk') as f:
            line_list = f.readlines()
    for line in line_list:
        yield line.strip()


def write_file(filename, data):
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(data)


def write_json(filename, doc, mode='a'):
    with open(filename, mode, encoding='utf-8') as f:
        f.write(str(json.dumps(doc, ensure_ascii=False)) + '\n')
